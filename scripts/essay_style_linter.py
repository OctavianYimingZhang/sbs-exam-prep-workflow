#!/usr/bin/env python3
"""Lint default workbook KP explanations for essay-style synthesis quality.

The linter protects the separation between internal source coverage and
student-facing prose. It does not judge biological correctness. It flags
page/slide narration, meta-writing instructions, OCR debris, and extreme
paragraph lengths so the workbook generator can rewrite before delivery.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:  # pragma: no cover
    load_workbook = None


EXPLANATION_HEADERS = {
    "essay-style sequential explanation",
    "essay-style synthesis",
    "essay style synthesis",
    "essay-style kp synthesis",
    "kp essay synthesis",
    "student-facing explanation",
    "student facing explanation",
    "explanation",
}

PREP_HEADERS = {
    "exam-facing prep",
    "exam facing prep",
    "predicted practice question",
    "prediction / prep",
    "prediction",
    "prep",
}

BANNED_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("page_number_trace", re.compile(r"\bPage\s+\d+\b", re.I)),
    ("slide_number_trace", re.compile(r"\bSlides?\s+\d+\b", re.I)),
    ("covers_pages", re.compile(r"\bcovers pages\b", re.I)),
    ("pages_should_be_read_as", re.compile(r"\bPages?\s+\d+(?:\s*[-–]\s*\d+)?\s+should be read as\b", re.I)),
    ("slide_sequence", re.compile(r"\bslide sequence\b", re.I)),
    ("remaining_linked_pages", re.compile(r"\bremaining linked pages\b", re.I)),
    ("first_establishes", re.compile(r"\bfirst establishes\b", re.I)),
    ("then_develops", re.compile(r"\bthen develops\b", re.I)),
    ("then_closes", re.compile(r"\bthen closes\b", re.I)),
    ("first_part_sequence", re.compile(r"\bfirst part of the sequence\b", re.I)),
    ("later_pages_add", re.compile(r"\blater pages add\b", re.I)),
    ("in_an_essay_answer", re.compile(r"\bIn an essay answer\b", re.I)),
    ("use_these_pages", re.compile(r"\buse these pages\b", re.I)),
    ("turn_pages", re.compile(r"\bTurn pages\b", re.I)),
    ("central_idea_block", re.compile(r"\bcentral idea for this block\b", re.I)),
    ("central_examinable_idea_block", re.compile(r"\bcentral examinable idea in this knowledge block\b", re.I)),
    ("knowledge_block_meta", re.compile(r"\b(?:this|the) knowledge block\b", re.I)),
    ("lecture_develops_across_slides", re.compile(r"\blecture develops it across slides\b", re.I)),
]

HOW_TO_WRITE_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("should_be_written_as", re.compile(r"\bshould be written as\b", re.I)),
    ("should_be_understood_as", re.compile(r"\bshould be understood as\b", re.I)),
    ("best_written_as", re.compile(r"\bbest written as\b", re.I)),
    ("mainly_an_argument_about", re.compile(r"\b(?:is|are)\s+mainly an argument about\b", re.I)),
    ("in_an_essay_answer", re.compile(r"\bIn an essay answer\b", re.I)),
    ("use_these_pages", re.compile(r"\buse these pages\b", re.I)),
    ("turn_pages", re.compile(r"\bTurn pages\b", re.I)),
    ("write_this_as", re.compile(r"\bwrite this as\b", re.I)),
    ("student_should", re.compile(r"\bstudents? should\b", re.I)),
    ("essay_should", re.compile(r"\bthe essay should\b", re.I)),
]

PREP_BANNED_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    ("turn_pages", re.compile(r"\bTurn pages\b", re.I)),
    ("use_these_pages", re.compile(r"\buse these pages\b", re.I)),
    ("into_one_paragraph", re.compile(r"\binto one paragraph\b", re.I)),
]

POSITIVE_CONNECTORS = re.compile(
    r"\b(because|therefore|consequently|as a result|depends on|regulates|drives|converts|constrains|enables|explains|links|integrates)\b",
    re.I,
)

REPEATED_PAGE_TRACE = re.compile(r"\bPage\s+\d+\b.*\bPage\s+\d+\b.*\bPage\s+\d+\b", re.I | re.S)
REPEATED_SLIDE_TRACE = re.compile(r"\bSlide\s+\d+\b.*\bSlide\s+\d+\b.*\bSlide\s+\d+\b", re.I | re.S)
BULLET_LIKE = re.compile(r"(^|\n)\s*(?:[-*•]|\d+[.)])\s+")
OCR_DEBRIS = re.compile(r"(?:[A-Za-z]-\s+[A-Za-z]|[A-Za-z]\s+[A-Za-z]\s+[A-Za-z]\s+[A-Za-z]\s+[A-Za-z]|�|□|_{3,})")
WORD_RE = re.compile(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)?")


@dataclass
class TextRecord:
    source: str
    row: int | None
    column: str
    text: str
    expected: str | None = None
    case_id: str | None = None


def words(text: str) -> list[str]:
    return WORD_RE.findall(text)


def normalise_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def find_column(headers: Sequence[object], candidates: set[str]) -> int | None:
    normalised = [normalise_header(header) for header in headers]
    for idx, header in enumerate(normalised):
        if header in candidates:
            return idx + 1
    for idx, header in enumerate(normalised):
        if "essay" in header and ("synthesis" in header or "explanation" in header):
            return idx + 1
    return None


def workbook_records(path: Path) -> list[TextRecord]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to lint .xlsx workbooks")
    wb = load_workbook(path, data_only=True, read_only=True)
    if "Exam_Prep_Map" not in wb.sheetnames:
        raise RuntimeError(f"{path} does not contain an Exam_Prep_Map sheet")
    ws = wb["Exam_Prep_Map"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    explanation_col = find_column(headers, EXPLANATION_HEADERS)
    if explanation_col is None:
        raise RuntimeError(f"{path} has no recognised essay-style explanation column")
    prep_col = find_column(headers, PREP_HEADERS)
    records: list[TextRecord] = []
    for row in ws.iter_rows(min_row=2):
        explanation = row[explanation_col - 1].value if len(row) >= explanation_col else None
        if explanation and str(explanation).strip():
            records.append(TextRecord(str(path), row[0].row, "explanation", str(explanation)))
        if prep_col is not None and len(row) >= prep_col:
            prep = row[prep_col - 1].value
            if prep and str(prep).strip():
                records.append(TextRecord(str(path), row[0].row, "prep", str(prep)))
    return records


def text_records(path: Path) -> list[TextRecord]:
    raw = path.read_text()
    paragraphs = [part.strip() for part in re.split(r"\n\s*\n", raw) if part.strip()]
    return [TextRecord(str(path), idx + 1, "explanation", paragraph) for idx, paragraph in enumerate(paragraphs)]


def fixture_records(path: Path) -> list[TextRecord]:
    data = json.loads(path.read_text())
    cases = data.get("cases", data if isinstance(data, list) else [])
    records = []
    for idx, case in enumerate(cases, start=1):
        records.append(
            TextRecord(
                source=str(path),
                row=idx,
                column=case.get("column", "explanation"),
                text=case.get("text", ""),
                expected=case.get("expected"),
                case_id=case.get("id", f"case_{idx}"),
            )
        )
    return records


def pattern_hits(text: str, patterns: Iterable[tuple[str, re.Pattern[str]]]) -> list[dict]:
    hits = []
    for name, pattern in patterns:
        for match in pattern.finditer(text):
            hits.append({"pattern": name, "phrase": match.group(0)})
            break
    return hits


def lint_records(records: list[TextRecord], min_words: int, max_words: int) -> dict:
    explanation_records = [record for record in records if record.column == "explanation"]
    prep_records = [record for record in records if record.column == "prep"]

    banned_rows = []
    repeated_trace_rows = []
    how_to_rows = []
    prep_banned_rows = []
    warnings = []

    for record in explanation_records:
        text = record.text.strip()
        banned = pattern_hits(text, BANNED_PATTERNS)
        if banned:
            banned_rows.append(row_payload(record, banned))
        if REPEATED_PAGE_TRACE.search(text) or REPEATED_SLIDE_TRACE.search(text):
            repeated_trace_rows.append(row_payload(record, [{"pattern": "repeated_page_or_slide_trace", "phrase": "three_or_more_page_or_slide_mentions"}]))
        how_to = pattern_hits(text, HOW_TO_WRITE_PATTERNS)
        if how_to:
            how_to_rows.append(row_payload(record, how_to))
        count = len(words(text))
        warning_reasons = []
        if count < min_words:
            warning_reasons.append(f"under_{min_words}_words")
        if count > max_words:
            warning_reasons.append(f"over_{max_words}_words")
        if BULLET_LIKE.search(text):
            warning_reasons.append("bullet_like_list")
        if OCR_DEBRIS.search(text):
            warning_reasons.append("possible_ocr_debris")
        if not POSITIVE_CONNECTORS.search(text):
            warning_reasons.append("no_positive_causal_connector")
        if warning_reasons:
            warnings.append({**row_payload(record, []), "warnings": warning_reasons, "word_count": count})

    for record in prep_records:
        hits = pattern_hits(record.text, PREP_BANNED_PATTERNS)
        if hits:
            prep_banned_rows.append(row_payload(record, hits))

    explanation_count = len(explanation_records)
    banned_rate = len(banned_rows) / explanation_count if explanation_count else 0.0
    how_to_rate = len(how_to_rows) / explanation_count if explanation_count else 0.0
    fail_reasons = []
    if explanation_count == 0:
        fail_reasons.append("no_explanations_found")
    if banned_rate > 0.05:
        fail_reasons.append("banned_pattern_rate_over_5_percent")
    if repeated_trace_rows:
        fail_reasons.append("repeated_page_or_slide_narration")
    if how_to_rate > 0.10:
        fail_reasons.append("how_to_write_language_rate_over_10_percent")
    if prep_banned_rows:
        fail_reasons.append("prep_column_contains_turn_pages_instruction")

    return {
        "pass": not fail_reasons,
        "fail_reasons": fail_reasons,
        "counts": {
            "explanations": explanation_count,
            "prep_cells": len(prep_records),
            "banned_rows": len(banned_rows),
            "repeated_trace_rows": len(repeated_trace_rows),
            "how_to_rows": len(how_to_rows),
            "prep_banned_rows": len(prep_banned_rows),
            "warnings": len(warnings),
        },
        "rates": {
            "banned_pattern_rate": round(banned_rate, 4),
            "how_to_write_rate": round(how_to_rate, 4),
        },
        "offending_rows": {
            "banned_patterns": banned_rows,
            "repeated_page_or_slide_narration": repeated_trace_rows,
            "how_to_write_language": how_to_rows,
            "prep_banned_patterns": prep_banned_rows,
        },
        "warnings": warnings[:100],
    }


def row_payload(record: TextRecord, hits: list[dict]) -> dict:
    return {
        "source": record.source,
        "row": record.row,
        "column": record.column,
        "case_id": record.case_id,
        "hits": hits,
        "preview": record.text[:220].replace("\n", " "),
    }


def fixture_expectation_report(records: list[TextRecord], min_words: int, max_words: int) -> dict:
    case_results = []
    for record in records:
        result = lint_records([record], min_words, max_words)
        expected = (record.expected or "").lower() or None
        expected_pass = None if expected not in {"pass", "fail"} else expected == "pass"
        matched = expected_pass is None or result["pass"] == expected_pass
        case_results.append(
            {
                "case_id": record.case_id,
                "expected": expected,
                "actual_pass": result["pass"],
                "matched": matched,
                "fail_reasons": result["fail_reasons"],
                "counts": result["counts"],
            }
        )
    return {
        "pass": all(case["matched"] for case in case_results),
        "cases": case_results,
        "counts": {
            "cases": len(case_results),
            "matched": sum(1 for case in case_results if case["matched"]),
            "mismatched": sum(1 for case in case_results if not case["matched"]),
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Lint workbook KP explanations for essay-style synthesis.")
    parser.add_argument("--workbook", action="append", type=Path, default=[])
    parser.add_argument("--text", action="append", type=Path, default=[])
    parser.add_argument("--fixture", action="append", type=Path, default=[])
    parser.add_argument("--output", type=Path)
    parser.add_argument("--min-words", type=int, default=60)
    parser.add_argument("--max-words", type=int, default=260)
    args = parser.parse_args()

    reports = []
    overall_pass = True

    for path in args.workbook:
        try:
            records = workbook_records(path)
            report = lint_records(records, args.min_words, args.max_words)
        except Exception as exc:
            report = {"pass": False, "fail_reasons": ["read_error"], "error": str(exc)}
        report.update({"type": "workbook", "path": str(path)})
        reports.append(report)
        overall_pass = overall_pass and bool(report["pass"])

    for path in args.text:
        try:
            records = text_records(path)
            report = lint_records(records, args.min_words, args.max_words)
        except Exception as exc:
            report = {"pass": False, "fail_reasons": ["read_error"], "error": str(exc)}
        report.update({"type": "text", "path": str(path)})
        reports.append(report)
        overall_pass = overall_pass and bool(report["pass"])

    for path in args.fixture:
        try:
            records = fixture_records(path)
            report = fixture_expectation_report(records, args.min_words, args.max_words)
        except Exception as exc:
            report = {"pass": False, "fail_reasons": ["read_error"], "error": str(exc)}
        report.update({"type": "fixture", "path": str(path)})
        reports.append(report)
        overall_pass = overall_pass and bool(report["pass"])

    if not reports:
        reports.append({"type": "none", "pass": False, "fail_reasons": ["no_inputs"]})
        overall_pass = False

    output = {"pass": overall_pass, "reports": reports}
    text = json.dumps(output, indent=2)
    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(text)
    else:
        print(text)
    return 0 if overall_pass else 1


if __name__ == "__main__":
    raise SystemExit(main())
